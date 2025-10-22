"""
거시경제 데이터 엑셀 레포트 생성 코드
openpyxl을 사용하여 스타일링된 엑셀 레포트를 생성합니다.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
from datetime import datetime


class EconomicReportGenerator:
    """거시경제 엑셀 레포트 생성기"""

    def __init__(self, data_path):
        """
        Args:
            data_path: 변환된 데이터 CSV 파일 경로
        """
        self.data_path = data_path
        self.df = None
        self.wb = None

    def load_data(self):
        """데이터 로드"""
        print("📥 데이터 로딩 중...")
        self.df = pd.read_csv(self.data_path)
        self.df['date'] = pd.to_datetime(self.df['date'])
        print(f"✅ {len(self.df)}개 레코드 로드 완료")
        return self.df

    def create_summary_sheet(self, ws):
        """요약 시트 생성"""
        print("📄 요약 시트 생성 중...")

        # 헤더
        ws['A1'] = 'Economic Indicators Report'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30

        # 보고서 정보
        ws['A3'] = 'Report Date:'
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'] = 'Data Period:'
        ws['B4'] = f"{self.df['date'].min().strftime('%Y-%m-%d')} ~ {self.df['date'].max().strftime('%Y-%m-%d')}"
        ws['A5'] = 'Total Records:'
        ws['B5'] = len(self.df)

        # 스타일 적용
        for row in range(3, 6):
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

        # 국가별 요약
        ws['A7'] = 'Summary by Country'
        ws['A7'].font = Font(size=12, bold=True)
        ws['A7'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws['A7'].font = Font(size=12, bold=True, color='FFFFFF')
        ws.merge_cells('A7:F7')

        country_summary = self.df.groupby('country').agg({
            'actual': 'count',
            'forecast_accuracy': 'mean'
        }).round(2)
        country_summary.columns = ['Data Points', 'Avg Forecast Accuracy (%)']

        start_row = 8
        ws['A8'] = 'Country'
        ws['B8'] = 'Data Points'
        ws['C8'] = 'Avg Forecast Accuracy (%)'

        for col in ['A8', 'B8', 'C8']:
            ws[col].font = Font(bold=True)
            ws[col].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        row = 9
        for country, data in country_summary.iterrows():
            ws[f'A{row}'] = country
            ws[f'B{row}'] = data['Data Points']
            ws[f'C{row}'] = data['Avg Forecast Accuracy (%)']
            row += 1

        # 열 너비 조정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25

        print("✅ 요약 시트 완료")

    def create_detail_sheet(self, ws):
        """상세 데이터 시트 생성"""
        print("📄 상세 데이터 시트 생성 중...")

        # 데이터 정렬
        df_sorted = self.df.sort_values(['date', 'country', 'indicator'], ascending=[False, True, True])

        # 컬럼 선택
        columns = ['date', 'country', 'indicator', 'forecast', 'actual', 'previous',
                   'forecast_error', 'forecast_error_pct', 'mom_change', 'mom_change_pct',
                   'forecast_accuracy', 'surprise', 'trend']
        df_display = df_sorted[columns].copy()

        # 날짜 포맷
        df_display['date'] = df_display['date'].dt.strftime('%Y-%m-%d')

        # 헤더 작성
        headers = ['Date', 'Country', 'Indicator', 'Forecast', 'Actual', 'Previous',
                   'Forecast Error', 'Error %', 'MoM Change', 'MoM %',
                   'Accuracy %', 'Surprise', 'Trend']

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 데이터 작성
        for r_idx, row in enumerate(dataframe_to_rows(df_display, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # 숫자 포맷
                if c_idx in [4, 5, 6, 7, 8, 9, 10, 11]:  # 숫자 컬럼
                    cell.number_format = '#,##0.00'

                # 조건부 서식
                if c_idx == 8:  # Forecast Error %
                    if isinstance(value, (int, float)):
                        if value > 0:
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        elif value < 0:
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # 테두리 추가
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=len(df_display) + 1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 필터 추가
        ws.auto_filter.ref = f"A1:{chr(64 + len(headers))}{len(df_display) + 1}"

        print(f"✅ 상세 데이터 시트 완료 ({len(df_display)} rows)")

    def create_indicator_sheets(self):
        """지표별 시트 생성"""
        print("📄 지표별 시트 생성 중...")

        indicators = self.df['indicator'].unique()

        for indicator in indicators:
            print(f"  - {indicator} 시트 생성 중...")
            ws = self.wb.create_sheet(title=indicator)

            # 해당 지표 데이터 필터링
            df_indicator = self.df[self.df['indicator'] == indicator].sort_values('date', ascending=False)

            # 컬럼 선택
            columns = ['date', 'country', 'forecast', 'actual', 'previous',
                       'forecast_error_pct', 'mom_change_pct', 'forecast_accuracy', 'trend']
            df_display = df_indicator[columns].copy()
            df_display['date'] = df_display['date'].dt.strftime('%Y-%m-%d')

            # 헤더
            headers = ['Date', 'Country', 'Forecast', 'Actual', 'Previous',
                       'Error %', 'MoM %', 'Accuracy %', 'Trend']

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 데이터 작성
            for r_idx, row in enumerate(dataframe_to_rows(df_display, index=False, header=False), 2):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    if c_idx in [3, 4, 5, 6, 7, 8]:
                        cell.number_format = '#,##0.00'

            # 열 너비 조정
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 15

        print(f"✅ {len(indicators)}개 지표별 시트 완료")

    def generate_report(self, output_path='Economic_Report.xlsx'):
        """전체 레포트 생성"""
        print("=" * 70)
        print("📊 엑셀 레포트 생성 시작")
        print("=" * 70)

        # 데이터 로드
        self.load_data()

        # 워크북 생성
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # 기본 시트 제거

        # 요약 시트
        summary_ws = self.wb.create_sheet(title='Summary')
        self.create_summary_sheet(summary_ws)

        # 상세 데이터 시트
        detail_ws = self.wb.create_sheet(title='All Data')
        self.create_detail_sheet(detail_ws)

        # 지표별 시트
        self.create_indicator_sheets()

        # 파일 저장
        print(f"\n💾 레포트 저장 중: {output_path}")
        self.wb.save(output_path)

        print()
        print("=" * 70)
        print(f"✅ 엑셀 레포트 생성 완료: {output_path}")
        print("=" * 70)

        return output_path


if __name__ == "__main__":
    # 레포트 생성 예제
    generator = EconomicReportGenerator('bloomberg_transformed_data.csv')
    report_path = generator.generate_report('Economic_Report.xlsx')

    print(f"\n📁 생성된 레포트: {report_path}")
